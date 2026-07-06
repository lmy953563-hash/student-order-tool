import streamlit as st
import pandas as pd
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="学生订购数据", layout="wide")
st.title("📊 学生订购数据智能分析看板")

# --- 专业报表生成模块 ---
def generate_perfect_excel(class_counts, total_sum, start_num, end_num):
    wb = Workbook()
    ws = wb.active
    dark_blue = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    steel_blue_40 = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    font_body = Font(name='微软雅黑', size=11)
    font_header = Font(name='微软雅黑', size=11, bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.column_dimensions['I'].width = 13.5; ws.column_dimensions['J'].width = 13.5
    ws.column_dimensions['K'].width = 1.5; ws.column_dimensions['L'].width = 13.5; ws.column_dimensions['M'].width = 13.5
    
    ws.merge_cells('I1:M1')
    ws['I1'] = f"截至 {datetime.now().strftime('%m月%d日%H:%M')} 开通情况"
    ws['I1'].fill = dark_blue; ws['I1'].font = Font(color="FFFFFF", bold=True, name='微软雅黑', size=14)
    ws['I1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36
    
    headers = ["班级", "开通人数", "", "班级", "开通人数"]
    for i, h in enumerate(headers, 9):
        cell = ws.cell(row=2, column=i, value=h)
        cell.fill = dark_blue if i == 11 else steel_blue_40
        cell.font = font_header; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = border
    
    total_classes = end_num - start_num + 1
    mid_point = (total_classes + 1) // 2
    for idx, i in enumerate(range(start_num, end_num + 1)):
        row_idx = (idx % mid_point) + 3
        col_start = 9 if idx < mid_point else 12
        cls_name = f"{i:02d}班" 
        ws.row_dimensions[row_idx].height = 24
        for c_idx in [col_start, col_start+1]:
            cell = ws.cell(row=row_idx, column=c_idx)
            cell.value = cls_name if c_idx == col_start else class_counts.get(f"{i}班", 0)
            cell.font = font_body; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = border
    
    last_row = mid_point + 2
    ws.merge_cells(f'K2:K{last_row}')
    for r in range(2, last_row + 1): ws.cell(row=r, column=11).fill = dark_blue
    ws.merge_cells(f'I{last_row+1}:M{last_row+1}')
    total_cell = ws.cell(last_row+1, 9, f"总计：{total_sum} 人")
    total_cell.fill = dark_blue; total_cell.font = Font(color="FFFFFF", bold=True, name='微软雅黑', size=12)
    total_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# --- 原有样式函数 ---
def apply_excel_style(ws):
    ws.column_dimensions['A'].width = 14; ws.column_dimensions['B'].width = 14; ws.column_dimensions['C'].width = 26.5 
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 20
        for cell in row:
            cell.font = Font(name='宋体', size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.fill = PatternFill(start_color='E6F0FF' if cell.row == 1 else 'FDF5E6', fill_type='solid')

# --- 【优化】添加缓存机制的数据加载函数 ---
@st.cache_data
def load_and_clean_data(file):
    df = pd.read_excel(file, sheet_name="学生订购信息汇总表")
    df = df.rename(columns={c: str(c).strip().replace('\ufeff', '') for c in df.columns})
    df = df.rename(columns={c: '姓名' for c in df.columns if '姓名' in c})
    df = df.rename(columns={c: '班级' for c in df.columns if '班级' in c})
    df = df.rename(columns={c: '学科' for c in df.columns if '学科' in c})
    
    df_clean = df.dropna(subset=['姓名', '班级', '学科']).copy()
    df_clean['学科'] = df_clean['学科'].astype(str)
    
    # 提前处理 exploded 数据
    df_exploded = df_clean.copy()
    df_exploded['学科'] = df_exploded['学科'].str.split('/')
    df_exploded = df_exploded.explode('学科')
    df_exploded['学科'] = df_exploded['学科'].str.strip()
    
    return df_clean, df_exploded

# --- 【优化】添加缓存的 Excel 生成函数 ---
@st.cache_data
def generate_class_excel(df_clean, selected_class):
    display_df = df_clean[df_clean['班级'] == selected_class].groupby(['班级', '姓名'])['学科'].first().reset_index()
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        display_df.to_excel(writer, index=False, sheet_name="导出")
        apply_excel_style(writer.sheets["导出"])
    return output.getvalue()

@st.cache_data
def generate_all_classes_excel(df_clean, all_classes):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for cls in all_classes:
            cls_data = df_clean[df_clean['班级'] == cls].groupby(['班级', '姓名'])['学科'].first().reset_index()
            # 限制 sheet name 长度在 31 字符内，防止 openpyxl 报错
            safe_cls_name = str(cls)[:31]
            cls_data.to_excel(writer, index=False, sheet_name=safe_cls_name)
            apply_excel_style(writer.sheets[safe_cls_name])
    return output.getvalue()

# --- 主程序 ---
uploaded_file = st.sidebar.file_uploader("请上传 Excel 文件", type=["xlsx"])

if uploaded_file is not None:
    try:
        # 使用缓存加载数据
        df_clean, df_exploded = load_and_clean_data(uploaded_file)

        # --- 还原原有功能：实时班级开通人数 (保留原文本框形式) ---
        st.subheader("✨ 实时班级开通人数")
        c_range = st.text_input("请输入统计范围 (如 1-36)：", "1-36")
        
        df_int = df_clean.copy()
        df_int['班级_int'] = df_int['班级'].astype(str).str.extract(r'(\d+)')[0].astype(int)
        cnts = df_int.groupby('班级_int')['姓名'].nunique().to_dict()
        total = sum(cnts.values())
        start, end = map(int, c_range.split('-'))
        excel_data = generate_perfect_excel({f"{k}班": v for k, v in cnts.items()}, total, start, end)
        
        st.download_button("💾 下载实时数据", excel_data, "实时班级统计.xlsx")
        st.markdown("---")
        
        # --- 原有功能 ---
        st.subheader("📈 核心统计数据")
        subj_totals = df_exploded.groupby('学科').size()
        cols = st.columns(2 + len(subj_totals))
        cols[0].metric("订购学生总人数", f"{len(df_clean['姓名'].unique())} 人")
        cols[1].metric("学科总订购数量", f"{len(df_exploded)} 科次")
        for i, (subj, count) in enumerate(subj_totals.items()): 
            cols[2 + i].metric(f"{subj} 总数", f"{count} 科")
        
        st.markdown("---")
        
        # --- 订购信息 ---
        st.subheader("🔍 订购信息")
        # 优化排序逻辑，避免非数字班级导致程序崩溃
        all_classes = sorted(df_clean['班级'].unique(), key=lambda x: int(re.search(r'\d+', x).group()) if re.search(r'\d+', x) else 999)
        selected_class = st.selectbox("🎯 按班级筛选显示：", all_classes)
        
        display_df = df_clean[df_clean['班级'] == selected_class].groupby(['班级', '姓名'])['学科'].first().reset_index()
        st.dataframe(display_df, use_container_width=True, hide_index=True)
        
        # 【优化】直接渲染下载按钮，完美解决原版代码中“点击导出后按钮消失/无法下载”的 Bug
        col1, col2 = st.columns([1, 4])
        with col1:
            single_class_excel = generate_class_excel(df_clean, selected_class)
            st.download_button("📥 导出选中班级", single_class_excel, f"{selected_class}_导出.xlsx")
            
        with col2:
            all_classes_excel = generate_all_classes_excel(df_clean, all_classes)
            st.download_button("📦 一键导出所有班级数据", all_classes_excel, "全部班级导出.xlsx")

        st.markdown("---")
        
        # --- 各班级学科统计表 ---
        st.subheader("📊 各班级学科统计表")
        pivot_df = df_exploded.groupby(['班级', '学科']).size().unstack(fill_value=0)
        pivot_df['总计'] = pivot_df.sum(axis=1)
        st.dataframe(pivot_df, use_container_width=True)

    except Exception as e:
        st.error(f"❌ 读取错误或数据格式异常: {e}")
